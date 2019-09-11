import openpyxl

wrkbk=input("Enter Workbook Name: ")
s=input("Enter Sheet Name: ")
wb=openpyxl.load_workbook(wrkbk)
sheet=wb[s]
col=int(input("Enter Product column number: "))
co=int(input("Enter column in which changes are to be made: "))
pname=input("Enter Product Name: ")
nval=input("Enter updated value: ")
nname=input("Enter name of new file: ")
if(nval.isdigit()):
    nval=int(nval)

try:
    for i in range(2,sheet.max_row+1):
        p=sheet.cell(row=i,column=col).value
        p=p.strip()
    
        if p==pname:
            sheet.cell(row=i,column=co).value=nval
            
    wb.save(nname)
    print("Success")
except:
    print("Something went wrong..")

        
